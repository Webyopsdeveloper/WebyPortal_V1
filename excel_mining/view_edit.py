import openpyxl

def edit_excel_file(file_path):
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    print("Current content of the Excel file:")
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end="\t")
        print()

    
    row_index = int(input("Enter the row index to edit (1-indexed): ")) - 1
    col_index = int(input("Enter the column index to edit (1-indexed): ")) - 1
    new_value = input("Enter the new value: ")

    
    sheet.cell(row=row_index + 1, column=col_index + 1).value = new_value

    
    workbook.save(file_path)
    print("Excel file updated successfully.")


excel_file_path = "sample_sheets/sheet1.xlsx"  


edit_excel_file(excel_file_path)
